from mysql.connector import Error
import mysql.connector
import openpyxl
archivo_excel = openpyxl.load_workbook('data1.xlsx')
hoja_excel = archivo_excel.active

filas = []
for fila in hoja_excel.iter_rows():
    fila_lista = [celda.value for celda in fila]
    filas.append(tuple(fila_lista))

modificar = ['City', 'Country']
indice = [i for i, nombre in enumerate(fila) if nombre in modificar]

for fila in hoja_excel.iter_rows():
    filtro = [celda.value for i, celda in enumerate(fila) if i not in indice]
    filas.append(tuple(filtro))


print(filas[2])


